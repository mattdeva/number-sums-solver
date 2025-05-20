import pandas as pd
import numpy as np
from openpyxl import load_workbook

from number_sums_solver.components.utils import _is_square_df

from typing import Sequence
    
def _df_unique_values(df:pd.DataFrame) -> list:
    return [v for v in np.unique(df.values) if v != df.iloc[0,0]]

def _get_colors_components(excel_path:str, sheet_name:str|None=None) -> tuple[pd.DataFrame, dict]:
    # assumes one sheet in excel for now (planning to change later)
    # i dont think theres a point to break up function even tho it does a couple things

    wb = load_workbook(excel_path)

    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    rows = []
    dict_ = {}
    for row in sheet.iter_rows():
        record = []
        for cell in row:
            if cell.fill and cell.fill.fgColor:
                color = cell.fill.fgColor.rgb[-6:]
                record.append(color)
            else:
                record.append(None)

            if isinstance(cell.value, tuple([float, int])):
                continue

            if '(' in cell.value:
                dict_[color] = int(cell.value.split('(')[-1].split(')')[0])
            
        rows.append(record)

    return pd.DataFrame(rows), dict_

class Colors:
    def __init__(self, col_df:pd.DataFrame, target_dict:dict[str, int]):
        self.col_df = col_df
        self.target_dict = target_dict
        self._check_input()

    def __len__(self):
        return len(self.target_dict)

    def _check_input(self):
        _is_square_df(self.col_df)

        df_values = _df_unique_values(self.col_df)
        dict_keys = list(self.target_dict)

        if not sorted(df_values) == sorted(dict_keys):
            raise ValueError(f'DataFrame and Target_Dict must have same values. got {df_values} and {dict_keys}')
        
    @classmethod
    def from_excel(cls, path:str, sheet_name:str|None=None):
        col_df, target_dict = _get_colors_components(path, sheet_name)
        return cls(col_df, target_dict) 
    
    @classmethod
    def blank(cls, size:int):
        lists = []
        for r in range(size+1):
            row = []
            for c in range(size+1):
                if r == 0 or c == 0:
                    row.append('000000')
                else:
                    row.append('white')
            lists.append(row)
        return cls(
            pd.DataFrame(lists),{'white':0}
        )

    @property
    def values(self):
        return list(self.target_dict)
    
    # NOTE: not a big fan of how i did the remapping.. what i did suggests there may be better way to organize things, but is good for now...
    def _remap_target_dict_from_sequence(self, sequence:Sequence) -> dict[str,int]:
        out_dict = {}
        # for i,j in zip(self.values, sequence): # not sure why this didnt work..
        for i,j in zip(list(self.target_dict), sequence):
            out_dict[j] = self.target_dict.pop(i)
        return out_dict

    def _remap_target_dict_from_dict(self, d:dict) -> dict[str, int]:
        out_dict = {}
        # for i,j in zip(self.values, d): # ^ same
        for i,j in zip(list(self.target_dict), d):
            out_dict[d[j]] = self.target_dict.pop(i)
        return out_dict
    
    def _remap_color_df_from_sequence(self, sequence:Sequence) -> None:
        col_values = _df_unique_values(self.col_df)
        for old_value, new_value in zip(col_values, sequence):
            self.col_df = self.col_df.replace(old_value, new_value)

    def _remap_color_df_from_dict(self, d:dict) -> None:
        for old_value, new_value in d.items():
            self.col_df = self.col_df.replace(old_value, new_value)
    
    def change_colors(self, input_:dict|Sequence) -> None:
        # a setter but also like not really
        if len(input_) != len(self):
            raise ValueError(f'length must match numbers of colors ({len(self)}. got {len(input_)})')
        
        if isinstance(input_, dict):
            self.target_dict = self._remap_target_dict_from_dict(input_)
            self._remap_color_df_from_dict(input_)
        elif isinstance(input_, Sequence):
            self.target_dict = self._remap_target_dict_from_sequence(input_)
            self._remap_color_df_from_sequence(input_)
        else:
            raise ValueError(f'input must be Sequence or dict. got {input_}')
        
        
